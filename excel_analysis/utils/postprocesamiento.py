import numpy as np
import pandas as pd
from openpyxl import load_workbook

from excel_analysis.utils.sistema_de_calificaciones import (
    assign_stock_grade,
)
from excel_analysis.models.neural_networks import obtener_threshold_optimo
from excel_analysis.constants import SheetResult, WEIGHTS


def calcular_calificaciones_y_umbral(df, y_pred, Y_test, price_column, sheet_name):
    # Asignar una calificación a la acción
    stock_grade = assign_stock_grade(df, y_pred, Y_test, price_column)
    predicted_return = np.sum(y_pred)

    # Obtener el umbral (threshold) óptimo
    optimal_threshold = obtener_threshold_optimo(Y_test, y_pred)
    conteo_positivos_reales = np.sum(Y_test)
    conteo_positivos_predichos = np.sum((y_pred > optimal_threshold).astype(int))

    final_value = conteo_positivos_reales - conteo_positivos_predichos

    return SheetResult(
        sheet_name,
        final_value,
        stock_grade,
        optimal_threshold,
        predicted_return,
        None,
        None,
    )


# Generación de resultados finales
def calificacion_a_numero(calificacion):
    mapeo = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
    return mapeo.get(calificacion, 0)


# Procesar el libro de Excel
def procesar_libro(nombre_archivo):
    libro = load_workbook(nombre_archivo)
    calificaciones_empresas = {}

    # Recorremos cada hoja y compilamos las calificaciones
    for nombre_hoja in WEIGHTS.keys():
        hoja = libro[nombre_hoja]
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            empresa = fila[0]
            calificacion = calificacion_a_numero(fila[-1])
            if empresa not in calificaciones_empresas:
                calificaciones_empresas[empresa] = {}
            calificaciones_empresas[empresa][nombre_hoja] = calificacion

    # Calcular la calificación promedio usando los pesos
    for empresa, calificaciones in calificaciones_empresas.items():
        total_calificacion_ponderada = sum(
            calificaciones[nombre_hoja] * WEIGHTS[nombre_hoja]
            for nombre_hoja in calificaciones
        )
        total_peso = sum(WEIGHTS[nombre_hoja] for nombre_hoja in calificaciones)
        calificaciones_empresas[empresa]["Promedio"] = (
            total_calificacion_ponderada / total_peso
        )

    return calificaciones_empresas


# Guardar los resultados en una nueva hoja
def guardar_resultados(nombre_archivo, calificaciones_empresas):
    with pd.ExcelWriter(nombre_archivo, mode="a", engine="openpyxl") as escritor:
        lista_calificaciones = []
        for empresa, calificaciones in calificaciones_empresas.items():
            fila = [empresa] + [
                calificaciones.get(nombre_hoja, "N/A") for nombre_hoja in WEIGHTS.keys()
            ]
            fila.append(calificaciones["Promedio"])
            lista_calificaciones.append(fila)

        columnas = ["Empresa"] + list(WEIGHTS.keys()) + ["Promedio"]
        df_resultados = pd.DataFrame(lista_calificaciones, columns=columnas)
        df_resultados.to_excel(escritor, sheet_name="Resultados", index=False)
