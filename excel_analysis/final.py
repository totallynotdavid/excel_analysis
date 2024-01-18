import pandas as pd
from openpyxl import load_workbook
from constants import weights


def grade_to_number(grade):
    mapping = {"A": 1, "B": 2, "C": 3, "D": 4, "E": 5, "F": 6}
    return mapping.get(grade, 0)


def process_workbook(file_name, weights):
    workbook = load_workbook(file_name)
    company_grades = {}

    for sheet_name in weights.keys():
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            company = row[0]
            grade = grade_to_number(
                row[-1]
            )  # La calificación está en la última columna
            if company not in company_grades:
                company_grades[company] = {}
            company_grades[company][sheet_name] = grade

    # Calculate the average grade using the weights
    for company, grades in company_grades.items():
        total_weighted_grade = sum(
            grades[sheet_name] * weights[sheet_name] for sheet_name in grades
        )
        total_weight = sum(weights[sheet_name] for sheet_name in grades)
        company_grades[company]["Average"] = total_weighted_grade / total_weight

    return company_grades


def save_results(file_name, company_grades, weights):
    with pd.ExcelWriter(file_name, mode="a", engine="openpyxl") as writer:
        # Create a DataFrame from the processed data
        grades_list = []
        for company, grades in company_grades.items():
            row = [company] + [grades.get(sheet, "N/A") for sheet in weights.keys()]
            row.append(grades["Average"])
            grades_list.append(row)

        columns = ["Company"] + list(weights.keys()) + ["Average Grade"]
        results_df = pd.DataFrame(grades_list, columns=columns)
        results_df.to_excel(writer, sheet_name="Results", index=False)


file_name = "resultados.xlsx"
company_grades = process_workbook(file_name, weights)
save_results(file_name, company_grades, weights)
